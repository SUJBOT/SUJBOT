"""Extract test dataset from criteria.xlsx and vystrizky.xlsx.

Uses an LLM to naturally merge criterion questions with snippet content,
producing fluent Czech questions. Also uses the LLM to create negated variants.

Usage:
    uv run python scripts/extract_criteria_dataset.py
    uv run python scripts/extract_criteria_dataset.py --model gpt-4o-mini
    uv run python scripts/extract_criteria_dataset.py --negate-ratio 0.3 --batch-size 5
    uv run python scripts/extract_criteria_dataset.py --output dataset/criteria_test.json
"""

import argparse
import asyncio
import json
import random
import re
from pathlib import Path

import openpyxl
from dotenv import load_dotenv
from litellm import acompletion


load_dotenv()

MERGE_SYSTEM_PROMPT = """\
Jsi expert na tvorbu evaluačních otázek pro hodnocení bezpečnostních zpráv jaderných zařízení.

Tvým úkolem je spojit kritérium (otázku) s výstřižkem (obsahem dokumentu) do jedné přirozené, \
plynulé české otázky. Výstřižek popisuje konkrétní obsah, který má bezpečnostní zpráva obsahovat — \
vlož ho do otázky tak, aby zněla přirozeně jako jedna věta/odstavec.

Pravidla:
- Zachovej odbornou terminologii beze změny
- Výsledek musí být JEDNA plynulá otázka (ne dvě oddělené)
- Nezkracuj důležité detaily z výstřižku
- Pokud je výstřižek dlouhý, zachovej hlavní body ale formuluj stručněji
- Otázka musí končit otazníkem
- NEODPOVÍDEJ na otázku, pouze ji přeformuluj
"""

NEGATE_SYSTEM_PROMPT = """\
Jsi expert na češtinu. Tvým úkolem je přeformulovat otázku do negativního smyslu — \
tak, aby správná odpověď na ni byla "ne".

Pravidla:
- Zachovej odbornou terminologii
- Změň smysl otázky (např. "Obsahuje..." → "Neobsahuje...", "Je splněn..." → "Není splněn...")
- Zachovej celý zbytek otázky beze změny
- Výsledek musí být otázka končící otazníkem
- Vrať POUZE přeformulovanou otázku, nic jiného
"""


def load_snippets(path: Path) -> dict[int, str]:
    """Load snippet ID -> text mapping from vystrizky.xlsx."""
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    snippets: dict[int, str] = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        snippet_id = row[2]
        snippet_text = row[5]
        if snippet_id is not None and snippet_text is not None:
            snippets[int(snippet_id)] = snippet_text.strip()
    wb.close()
    return snippets


def parse_snippet_ids(vystrizky_str: str | None) -> list[int]:
    """Parse pipe-delimited snippet IDs like '|6|;|74|;|377|' into [6, 74, 377]."""
    if not vystrizky_str:
        return []
    return [int(m) for m in re.findall(r"\|(\d+)\|", vystrizky_str)]


def extract_main_question(text: str) -> str:
    """Extract the main question part, stripping guidelines and extra whitespace."""
    parts = re.split(r"\n\s*Guideline:", text, maxsplit=1, flags=re.IGNORECASE)
    question = parts[0].strip()
    question = re.sub(r"\n{2,}", "\n", question)
    question = re.sub(r"[ \t]+", " ", question)
    question = question.replace("\xa0", " ")
    return question.strip()


def extract_criteria(path: Path) -> list[dict]:
    """Extract criteria from criteria.xlsx."""
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    criteria = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        criterion_id = row[1]
        if criterion_id is None:
            continue
        criteria.append(
            {
                "id": int(criterion_id),
                "phase": (row[2] or "").strip(),
                "group": (row[3] or "").strip(),
                "text": (row[4] or "").strip(),
                "binding": (row[5] or "").strip(),
                "status": (row[6] or "").strip(),
                "snippet_ids_raw": row[7] or "",
                "chapter_structure": (row[8] or "").strip(),
            }
        )
    wb.close()
    return criteria


async def llm_merge(question: str, snippet: str, model: str) -> str:
    """Use LLM to naturally merge a criterion question with a snippet."""
    user_msg = (
        f"Kritérium (otázka):\n{question}\n\n"
        f"Výstřižek (obsah):\n{snippet}\n\n"
        f"Spoj je do jedné přirozené otázky:"
    )
    resp = await acompletion(
        model=model,
        messages=[
            {"role": "system", "content": MERGE_SYSTEM_PROMPT},
            {"role": "user", "content": user_msg},
        ],
        temperature=0.3,
        max_tokens=1024,
    )
    return resp.choices[0].message.content.strip()


async def llm_negate(question: str, model: str) -> str:
    """Use LLM to create a negated version of a question."""
    resp = await acompletion(
        model=model,
        messages=[
            {"role": "system", "content": NEGATE_SYSTEM_PROMPT},
            {"role": "user", "content": question},
        ],
        temperature=0.2,
        max_tokens=1024,
    )
    return resp.choices[0].message.content.strip()


async def process_entry(
    criterion: dict,
    snippet_id: int,
    snippet_text: str,
    model: str,
    should_negate: bool,
    semaphore: asyncio.Semaphore,
) -> list[dict]:
    """Process one criterion+snippet pair: merge, and optionally negate."""
    main_question = extract_main_question(criterion["text"])
    # Truncate very long snippets
    snip = snippet_text[:800] if len(snippet_text) > 800 else snippet_text

    async with semaphore:
        merged = await llm_merge(main_question, snip, model)

    base = {
        "criterion_id": criterion["id"],
        "snippet_id": snippet_id,
        "phase": criterion["phase"],
        "group": criterion["group"],
        "binding": criterion["binding"],
        "chapter_structure": criterion["chapter_structure"],
    }

    results = [
        {**base, "question": merged, "expected_answer": "yes", "is_negated": False}
    ]

    if should_negate:
        async with semaphore:
            negated = await llm_negate(merged, model)
        results.append(
            {**base, "question": negated, "expected_answer": "no", "is_negated": True}
        )

    return results


async def build_dataset(
    criteria: list[dict],
    snippets: dict[int, str],
    model: str,
    negate_ratio: float = 0.25,
    concurrency: int = 10,
    seed: int = 42,
) -> list[dict]:
    """Build test dataset using LLM to merge questions with snippets."""
    rng = random.Random(seed)
    semaphore = asyncio.Semaphore(concurrency)

    # Build all tasks
    tasks = []
    no_snippet_entries = []

    for criterion in criteria:
        if not criterion["text"]:
            continue
        snippet_ids = parse_snippet_ids(criterion["snippet_ids_raw"])
        resolved = [(sid, snippets[sid]) for sid in snippet_ids if sid in snippets]

        if not resolved:
            main_q = extract_main_question(criterion["text"])
            no_snippet_entries.append(
                {
                    "criterion_id": criterion["id"],
                    "phase": criterion["phase"],
                    "group": criterion["group"],
                    "binding": criterion["binding"],
                    "chapter_structure": criterion["chapter_structure"],
                    "snippet_ids": snippet_ids,
                    "question": main_q,
                    "expected_answer": "yes",
                    "is_negated": False,
                }
            )
            continue

        for sid, stxt in resolved:
            should_negate = rng.random() < negate_ratio
            tasks.append(
                process_entry(criterion, sid, stxt, model, should_negate, semaphore)
            )

    total = len(tasks)
    print(f"  {total} LLM merge tasks + {len(no_snippet_entries)} no-snippet entries")

    # Process with progress reporting
    dataset = list(no_snippet_entries)
    done = 0
    for coro in asyncio.as_completed(tasks):
        entries = await coro
        dataset.extend(entries)
        done += 1
        if done % 50 == 0 or done == total:
            print(f"  Progress: {done}/{total} ({done * 100 // total}%)")

    return dataset


def main():
    parser = argparse.ArgumentParser(description="Extract criteria test dataset (LLM-powered)")
    parser.add_argument(
        "--criteria", type=Path, default=Path("extraction_dataset/criteria.xlsx"),
    )
    parser.add_argument(
        "--snippets", type=Path, default=Path("extraction_dataset/vystrizky.xlsx"),
    )
    parser.add_argument(
        "--output", type=Path, default=Path("dataset/criteria_eval.json"),
    )
    parser.add_argument(
        "--model", type=str, default="gpt-4o-mini",
        help="LLM model for merging (default: gpt-4o-mini)",
    )
    parser.add_argument("--negate-ratio", type=float, default=0.25)
    parser.add_argument("--concurrency", type=int, default=10)
    parser.add_argument("--seed", type=int, default=42)
    parser.add_argument(
        "--limit", type=int, default=0,
        help="Limit number of criteria to process (0 = all)",
    )
    args = parser.parse_args()

    print(f"Loading snippets from {args.snippets}...")
    snippets = load_snippets(args.snippets)
    print(f"  Loaded {len(snippets)} snippets")

    print(f"Loading criteria from {args.criteria}...")
    criteria = extract_criteria(args.criteria)
    print(f"  Loaded {len(criteria)} criteria")

    if args.limit > 0:
        criteria = criteria[: args.limit]
        print(f"  Limited to first {args.limit} criteria")

    print(f"Building dataset (model={args.model}, negate_ratio={args.negate_ratio})...")
    dataset = asyncio.run(
        build_dataset(
            criteria, snippets, args.model, args.negate_ratio, args.concurrency, args.seed,
        )
    )

    positive = sum(1 for d in dataset if not d["is_negated"])
    negated = sum(1 for d in dataset if d["is_negated"])
    print(f"\n  {positive} positive + {negated} negated = {len(dataset)} total entries")

    # Show a few examples
    print("\n--- Sample merged questions ---")
    for entry in dataset[:4]:
        q = entry["question"]
        preview = q[:250] + "..." if len(q) > 250 else q
        tag = "NEG" if entry["is_negated"] else "POS"
        print(f"  [{tag}] (crit={entry['criterion_id']}) {preview}\n")

    args.output.parent.mkdir(parents=True, exist_ok=True)
    with open(args.output, "w", encoding="utf-8") as f:
        json.dump(
            {
                "description": "Criteria evaluation dataset - LLM-merged questions with snippet content",
                "model_used": args.model,
                "source_criteria": str(args.criteria),
                "source_snippets": str(args.snippets),
                "total_entries": len(dataset),
                "positive_entries": positive,
                "negated_entries": negated,
                "entries": dataset,
            },
            f,
            ensure_ascii=False,
            indent=2,
        )

    print(f"Written to {args.output}")


if __name__ == "__main__":
    main()
